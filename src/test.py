import openpyxl

# 特定の列を検索
def search_column(column, keyword):
    result = []
    for cell in column:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if value == keyword:
            cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
            result.append(cell_address)

    return result

# 特定の行を検索
def search_row(row, keyword):
    result = []
    for cell in row:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if value == keyword:
            cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
            result.append(cell_address)
            
    return result

# 特定の範囲を検索
def search_rectangle(rectangle, keyword):
    result = []
    for col in rectangle:
        for cell in col:
            # セルのデータを文字列に変換
            try:
                value = str(cell.value)
            # 文字列に変換できないデータはスキップ
            except:
                continue
            # キーワードに一致するセルの番地を取得
            if value == keyword:
                cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
                result.append(cell_address)
            
    return result

# シート全体を検索
def search_entire_sheet(ws, keyword):
    result = []
    for col in ws.columns:
        for cell in col:
            # セルのデータを文字列に変換
            try:
                value = str(cell.value)
            # 文字列に変換できないデータはスキップ
            except:
                continue
            # キーワードに一致するセルの番地を取得
            if value == keyword:
                cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
                result.append(cell_address)
            
    return result

def main():
    filename = 'foo.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']

    # result = search_column(ws['A'], 'scarf')
    # result = search_row(ws['42'], 'neutral')
    result = search_rectangle(ws['B5':'F38'], 'whisky')
    # result = search_entire_sheet(ws, 'soda_bottle')
    print(result)

if __name__ == '__main__':
    main()